import csv
import os

from openpyxl import load_workbook


def load_csv(file_path):

    with open(
        file_path,
        "r",
        encoding="utf-8-sig",
        newline=""
    ) as f:

        reader = csv.DictReader(f)

        return list(reader)


def load_xlsx(file_path):

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )

    data = {}

    for sheet in workbook.worksheets:

        rows = list(
            sheet.iter_rows(values_only=True)
        )

        if not rows:
            data[sheet.title] = []
            continue

        headers = []

        for index, value in enumerate(rows[0]):

            if value is None:
                headers.append(f"column_{index + 1}")
            else:
                headers.append(str(value))

        sheet_data = []

        for row in rows[1:]:

            item = {}

            for index, header in enumerate(headers):

                value = row[index] if index < len(row) else None

                item[header] = value

            sheet_data.append(item)

        data[sheet.title] = sheet_data

    workbook.close()

    return data


def load_data(file_path):

    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".csv":

        return load_csv(file_path)

    if extension == ".xlsx":

        return load_xlsx(file_path)

    raise ValueError(
        "Chỉ hỗ trợ CSV và XLSX."
    )


def get_numeric_columns(rows):

    if not rows:
        return []

    columns = []

    for column in rows[0].keys():

        values = []

        for row in rows:

            value = row.get(column)

            if isinstance(value, (int, float)):

                values.append(value)

        if values and len(values) == len(rows):

            columns.append(column)

    return columns


def analyze_rows(rows):

    if not rows:
        return {
            "rows": 0,
            "columns": []
        }

    numeric_columns = get_numeric_columns(rows)

    result = {
        "rows": len(rows),
        "columns": list(rows[0].keys()),
        "numeric_columns": numeric_columns,
        "statistics": {}
    }

    for column in numeric_columns:

        values = [
            row[column]
            for row in rows
        ]

        min_value = min(values)
        max_value = max(values)

        min_row = next(
            row
            for row in rows
            if row[column] == min_value
        )

        max_row = next(
            row
            for row in rows
            if row[column] == max_value
        )

        result["statistics"][column] = {
            "sum": sum(values),
            "average": sum(values) / len(values),
            "min": min_value,
            "min_row": min_row,
            "max": max_value,
            "max_row": max_row
        }

    return result


def analyze_file(file_path):

    data = load_data(file_path)

    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".csv":

        return analyze_rows(data)

    if extension == ".xlsx":

        result = {}

        for sheet_name, rows in data.items():

            result[sheet_name] = analyze_rows(rows)

        return result
