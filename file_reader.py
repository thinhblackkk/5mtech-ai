import os
import csv
import json

import pymupdf
from docx import Document
from openpyxl import load_workbook


def read_text_file(file_path):

    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def read_pdf_file(file_path):

    pdf = pymupdf.open(file_path)

    pages = []

    for page in pdf:
        pages.append(page.get_text())

    pdf.close()

    return "\n\n".join(pages)


def read_docx_file(file_path):

    document = Document(file_path)

    paragraphs = []

    for paragraph in document.paragraphs:

        text = paragraph.text.strip()

        if text:
            paragraphs.append(text)

    return "\n\n".join(paragraphs)


def read_csv_file(file_path):

    rows = []

    with open(
        file_path,
        "r",
        encoding="utf-8-sig",
        newline=""
    ) as f:

        reader = csv.reader(f)

        for row in reader:

            rows.append(" | ".join(row))

    return "\n".join(rows)


def read_json_file(file_path):

    with open(
        file_path,
        "r",
        encoding="utf-8"
    ) as f:

        data = json.load(f)

    return json.dumps(
        data,
        ensure_ascii=False,
        indent=2
    )

def read_xlsx_file(file_path):

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )

    sheets = []

    for sheet in workbook.worksheets:

        rows = []

        for row in sheet.iter_rows(values_only=True):

            values = []

            for value in row:

                if value is None:
                    values.append("")
                else:
                    values.append(str(value))

            rows.append(" | ".join(values))

        sheet_text = f"[Sheet: {sheet.title}]\n" + "\n".join(rows)

        sheets.append(sheet_text)

    workbook.close()

    return "\n\n".join(sheets)


def read_file(file_path):

    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".pdf":

        return read_pdf_file(file_path)

    if extension == ".docx":

        return read_docx_file(file_path)

    if extension == ".csv":

        return read_csv_file(file_path)

    if extension == ".json":

        return read_json_file(file_path)

    if extension == ".xlsx":

        return read_xlsx_file(file_path)

    return read_text_file(file_path)